import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
import shutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, GradientFill
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel import *


def autofit_column_and_row(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    for column_cells in sheet.iter_cols(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    for row_cells in sheet.rows:
        max_height = 0
        for cell in row_cells:
            try:
                lines = str(cell.value).count("\n") + 1
                height = (lines * 12) + 4
                if height > max_height:
                    max_height = height
            except:
                pass
        sheet.row_dimensions[row_cells[0].row].height = max_height


if __name__ == "__main__":
    # 파일 로드
    workbook = load_workbook('/Users/kimjunghoo/Desktop/uipath_python/연습용.xlsx')
    sheet = workbook['Sheet1']

    # 열 너비와 행 높이를 자동으로 조정
    autofit_column_and_row(sheet)

    # 변경된 내용을 저장
    workbook.save('/Users/kimjunghoo/Desktop/uipath_python/연습용.xlsx')

from openpyxl import load_workbook

def autofit_column_and_row(sheet, start_row, end_row, start_column, end_column):
    for column_cells in sheet.iter_cols(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    for row_cells in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        max_height = 0
        for cell in row_cells:
            try:
                lines = str(cell.value).count("\n") + 1
                height = (lines * 12) + 4
                if height > max_height:
                    max_height = height
            except:
                pass
        sheet.row_dimensions[row_cells[0].row].height = max_height

if __name__ == "__main__":
    workbook = load_workbook('example.xlsx')
    sheet = workbook.active
    autofit_column_and_row(sheet, start_row=2, end_row=3, start_column=1, end_column=2)
    workbook.save('example_auto_fitted.xlsx')





