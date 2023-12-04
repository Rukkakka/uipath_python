from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import re
import numpy as np
from typing import List, Tuple


def read_cell(file_name: str, sheet_name: [str, int], cell: str, only_data: bool = True):
    wb = load_workbook(file_name, data_only=only_data)
    ws = wb[sheet_name]
    cell_value = ws[cell].value
    return cell_value


def read_range(file_name: str, sheet_name: [str, int], range: str = None, header_y_n: bool = True,
               only_data: bool = True):
    # 시트 호출
    wb = load_workbook(file_name, data_only=only_data)
    ws = wb[sheet_name]

    # 데이터 초기화
    data = []

    # 시트/범위 분류
    # 시트 호출
    if range == None:
        for row in ws.iter_rows(values_only=True):
            data.append(row)

    # 범위 호출
    else:
        pattern = r'[a-z A-Z]+'
        result = re.findall(pattern, range)
        alphabet = result
        alphabet_count = len(alphabet)

        pattern = r'\d+'
        result = re.findall(pattern, range)
        numbers = result
        numbers_count = len(numbers)

        if alphabet_count == 1 and numbers_count == 1:
            # A1
            start_column = alphabet[0]
            start_column = column_index_from_string(start_column)
            start_row = int(numbers[0])

            for row in ws.iter_rows(min_row=start_row, min_col=start_column):
                row_data = [cell.value for cell in row]
                data.append(row_data)

        elif alphabet_count == 2 and numbers_count == 2:
            # A1:B2
            start_column = alphabet[0]
            start_column = column_index_from_string(start_column)
            start_row = int(numbers[0])

            end_column = alphabet[1]
            end_column = column_index_from_string(end_column)
            end_row = int(numbers[1])

            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
                row_data = [cell.value for cell in row]
                data.append(row_data)

    if header_y_n:
        # 리스트를 DataFrame으로 변환
        columns = data[0]  # 첫 번째 행을 컬럼으로 사용
        data = data[1:]  # 첫 번째 행을 제외한 나머지 데이터
        df = pd.DataFrame(data, columns=columns)
    else:
        df = pd.DataFrame(data)

    return df


def data_input(file_name: str, sheet_name: [str, int], start_cell: str, end_cell: str, data: str, fluctuate=0):
    """
    start_cell 및 end_cell cell값으로 기재 ex. A1, B3 등등..
    만약 최대 행
    갯수까지 라인을 그리고 싶다면 end_cell은 컬럼명+MAX로 한다. ex(B+MAX)
    """
    wb = load_workbook(file_name, data_only=False)
    ws = wb[sheet_name]

    if end_cell in 'MAX' or 'max':

        # cell값에서 column값 추출
        pattern = r'^([A-Z]+)'
        result = re.findall(pattern, start_cell)
        start_column = result[0]

        pattern = r'\d+'
        result = re.findall(pattern, start_cell)
        start_row = int(result[0])

        end_column = end_cell.split("+")[0]

        for col in range(ord(start_column), ord(end_column) + 1):

            column_letter = chr(col)

            for row in range(start_row, ws.max_row + 1):

                if '수웃자' in data:
                    data_modify = data.replace("수웃자", str(row + fluctuate))

                else:
                    data_modify = data
                ws[f'{column_letter}{str(row)}'] = data_modify

    else:

        # cell값에서 column값 추출
        pattern = r'^([A-Z]+)'
        result = re.findall(pattern, start_cell)
        start_column = result[0]

        result = re.findall(pattern, end_cell)
        end_column = result[0]

        # cell값에서 row값 추출
        pattern = r'\d+'
        result = re.findall(pattern, start_cell)
        start_row = int(result[0])

        result = re.findall(pattern, end_cell)
        end_row = int(result[0])

        for col in range(ord(start_column), ord(end_column) + 1):

            column_letter = chr(col)

            for row in range(start_row, end_row + 1):
                ws[f'{column_letter}{str(row)}'] = data

    wb.save(file_name)  # 값 저장


def write_cell(file_name: str, sheet_name: [str, int], cell: str, data: any, sheet_create: bool = True):
    wb = load_workbook(file_name, data_only=False)  # 파일 호출
    try:
        ws = wb[sheet_name]  # 시트호출
    except:
        if sheet_create:
            wb.create_sheet(sheet_name)
            print(f'{sheet_name} 생성')
            ws = wb[sheet_name]
        else:
            raise Exception(f'{sheet_name} 없습니다.')

    ws[cell] = data
    wb.save(file_name)  # 값 저장


def write_range(file_name: str, df, sheet_name: [str, int], cell: str = None, headers: bool = False,
                index: bool = False, sheet_create: bool = True):
    # cell값이 지정이 되어 있지 않다면 항상 A1부터 기입되게 설정
    if cell is None:
        cell = 'A1'

    pattern = r'[a-z A-Z]+'
    result = re.findall(pattern, cell)
    start_column = result[0]
    start_column = column_index_from_string(start_column) - 1

    pattern = r'\d+'
    result = re.findall(pattern, cell)
    start_row = int(result[0])

    wb = load_workbook(file_name, data_only=False)  # 파일 호출
    try:
        ws = wb[sheet_name]  # 시트호출
    except:
        if sheet_create:
            wb.create_sheet(sheet_name)
            print(f'{sheet_name} 생성')
            ws = wb[sheet_name]
        else:
            raise Exception(f'{sheet_name} 없습니다.')

    # 데이터프레임을 엑셀에 추가
    rows = dataframe_to_rows(df, index, headers)

    for row in rows:
        for idx, cell_value in enumerate(row, start=1):
            ws.cell(row=start_row, column=start_column + idx, value=cell_value)
        start_row += 1

    wb.save(file_name)  # 값 저장


def append_range(save_file_name: str, save_sheet_name: [str, int], read_file_name: str, read_sheet_name: [str, int],
                 headers: bool = False):
    wb1 = load_workbook(save_file_name, data_only=False)
    ws1 = wb1[save_sheet_name]

    wb2 = load_workbook(read_file_name, data_only=False)
    ws2 = wb2[read_sheet_name]

    data = []

    if headers:
        min = 1
    else:
        min = 2

    for row in ws2.iter_rows(min_ro=min, values_only=True):
        data.append(row)

    # 데이터프레임 생성
    df = pd.DataFrame(data)

    for row in dataframe_to_rows(df, index=False, header=False):
        ws1.append(row)

    wb1.save(save_file_name)


def append_range_workbook(file_name: str, sheet_name: [str, int], df):
    wb = load_workbook(file_name, data_only=False)
    ws = wb[sheet_name]

    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    wb.save(file_name)


def line_builder(file_name: str, sheet_name: [str, int], start_cell: str, end_cell: str, line_type='thin'):
    """
    start_cell 및 end_cell cell값으로 기재 ex. A1, B3 등등..
    만약 최대 행 갯수까지 라인을 그리고 싶다면 end_cell은 컬럼명+MAX로 한다. ex(B+MAX)
    """

    border = Border(left=Side(border_style=line_type),
                    right=Side(border_style=line_type),
                    top=Side(border_style=line_type),
                    bottom=Side(border_style=line_type))

    wb = load_workbook(file_name, data_only=False)
    ws = wb[sheet_name]

    # for col in range(ord(start_column), ord(end_column) + 1):

    #     column_letter = chr(col)

    #     for row in range(1, ws.max_row + 1):

    #         cell = ws[f'{column_letter}{str(row)}']
    #         cell.border = border

    # wb.save(file_name) # 값 저장

    if end_cell in 'MAX' or 'max':

        # cell값에서 column값 추출
        pattern = r'^([A-Z]+)'
        result = re.findall(pattern, start_cell)
        start_column = result[0]

        end_column = end_cell.split("+")[0]

        for col in range(ord(start_column), ord(end_column) + 1):

            column_letter = chr(col)

            for row in range(1, ws.max_row + 1):
                cell = ws[f'{column_letter}{str(row)}']
                cell.border = border

    else:

        # cell값에서 column값 추출
        pattern = r'^([A-Z]+)'
        result = re.findall(pattern, start_cell)
        start_column = result[0]

        result = re.findall(pattern, end_cell)
        end_column = result[0]

        # cell값에서 row값 추출
        pattern = r'\d+'
        result = re.findall(pattern, start_cell)
        start_row = int(result[0])

        result = re.findall(pattern, end_cell)
        end_row = int(result[0])

        for col in range(ord(start_column), ord(end_column) + 1):

            column_letter = chr(col)

            for row in range(start_row, end_row + 1):
                cell = ws[f'{column_letter}{str(row)}']
                cell.border = border

    wb.save(file_name)  # 값 저장


def color_input(file_name: str, sheet_name: [str, int], range: str, color_code: str, type='solid'):
    wb = load_workbook(file_name, data_only=False)

    # 시트 선택
    ws = wb[sheet_name]

    """
    색깔 채우는 조건 설정
    color_code 원하는 색깔(색깔 코드로)
    fill_type 어떻게 채울지 
    """
    fill = PatternFill(start_color=color_code, end_color=color_code, fill_type=type)  # 색깔 채우는 type 선택

    # 원하는 범위 지정( column 전체 및 row 전체도 가능하고 범위도 가능)
    cell_range = ws[range]
    for row in cell_range:
        try:
            for cell in row:
                cell.fill = fill
        except:
            row.fill = fill

    wb.save(file_name)


def autofit_range(file_name: str, sheet_name: [str, int], column_y_n: bool = True, row_y_n: bool = True):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    if column_y_n:
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    if row_y_n:
        for row_cells in ws.rows:
            max_height = 0
            for cell in row_cells:
                try:
                    lines = str(cell.value).count("\n") + 1
                    height = (lines * 12) + 4
                    if height > max_height:
                        max_height = height
                except:
                    pass
            ws.row_dimensions[row_cells[0].row].height = max_height

    wb.save(file_name)


def clear_sheet_range_table(file_name: str, sheet_name: [str, int], header_y_n: bool = True, range: str = None):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    if range:
        pattern = r'[a-z A-Z]+'
        result = re.findall(pattern, range)
        alphabet = result
        start_column = alphabet[0]
        start_column = int(ord(start_column)) - 64
        end_column = alphabet[1]
        end_column = int(ord(end_column)) - 64

        pattern = r'\d+'
        result = re.findall(pattern, range)
        numbers = result
        start_row = int(numbers[0])
        end_row = int(numbers[1])

        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
            for cell in row:
                cell.value = None

    else:

        if header_y_n:
            ws.delete_rows(2, ws.max_row)
        else:
            ws.delete_rows(1, ws.max_row)

    wb.save(file_name)


def copy_paste_range(file_name1: str, sheet_name1: [str, int], file_name2: str, sheet_name2: [str, int],
                     r_range: str = None, w_range: str = None, header_y_n: bool = True, only_data: bool = True,
                     change: bool = False):
    if change:
        df = read_range(file_name1, sheet_name1, r_range, False, only_data)
        df = df.transpose()
        write_range(file_name2, df, sheet_name2, w_range, False)
    else:
        df = read_range(file_name1, sheet_name1, r_range, header_y_n, only_data)
        write_range(file_name2, df, sheet_name2, w_range, header_y_n)


def delete_column(file_name: str, sheet_name: [str, int], column: str):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    df = pd.read_excel(file_name, sheet_name=sheet_name)
    try:
        column_index = df.columns.get_loc(column) + 1
        ws.delete_cols(column_index)
        wb.save(file_name)

    except:
        pattern = r'[a-z A-Z]+'
        result = re.findall(pattern, column)
        alphabet = result
        alphabet_count = len(alphabet)
        if alphabet_count == 1:
            ws.delete_cols(column_index_from_string(alphabet[0]))
            wb.save(file_name)
        elif alphabet_count == 2 and ':' in column:
            start_column = alphabet[0]
            end_column = alphabet[1]
            delete_numbers = column_index_from_string(end_column) - column_index_from_string(start_column) + 1
            ws.delete_cols(column_index_from_string(start_column), amount=delete_numbers)
            wb.save(file_name)
        elif ',' in column:
            column_split_str = column.split(',')
            column_split_int = [column_index_from_string(i) for i in column_split_str]
            column_split_int = sorted(column_split_int, reverse=True)
            for i in column_split_int:
                ws.delete_cols(i)
                wb.save(file_name)


def delete_rows(file_name: str, sheet_name: [str, int], row: [str, int]):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    try:
        row_index = int(row)
        ws.delete_rows(idx=row_index)
        wb.save(file_name)
    except:
        if '-' in row:
            pattern = r'\d+'
            result = re.findall(pattern, row)
            numbers = result
            row_index = int(numbers[1]) - int(numbers[0]) + 1
            for i in range(row_index):
                ws.delete_rows(idx=int(numbers[0]))
                wb.save(file_name)
        elif ',' in row:
            pattern = r'\d+'
            result = re.findall(pattern, row)
            numbers = [int(i) for i in result]
            numbers = sorted(numbers, reverse=True)
            for i in numbers:
                ws.delete_rows(idx=i)
                wb.save(file_name)


def export_to_csv(file_name: str, sheet_name: [str, int], csv_file_name: str):
    df = pd.read_excel(file_name, sheet_name=sheet_name, engine='openpyxl')
    df.to_csv(csv_file_name, index=False, encoding='utf-8-sig')


def fill_range(file_name: str, sheet_name: [str, int], range_value: str, value: [str, int], value_change: bool = True):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    pattern = r'[a-z A-Z]+'
    column_result = re.findall(pattern, range_value)
    start_column = column_result[0]
    start_column = column_index_from_string(start_column)
    end_column = column_result[1]
    end_column = column_index_from_string(end_column)

    pattern = r'\d+'
    row_result = re.findall(pattern, range_value)
    start_row = int(row_result[0])
    end_row = int(row_result[1])

    if value_change:
        if type(value) == int:
            for row_idx, row in enumerate(range(start_row, end_row + 1)):
                for col_idx, col in enumerate(range(start_column, end_column + 1)):
                    new_value = value + col_idx
                    ws.cell(row=row, column=col, value=new_value)
                value = value + 1
        else:
            if '=' in value:
                pattern = r'[A-Z]+\d+'
                value_result = re.findall(pattern, value)
                dict = {}

                for row_idx, row in enumerate(range(start_row, end_row + 1)):
                    for col_idx, col in enumerate(range(start_column, end_column + 1)):
                        for i in value_result:
                            cell_colume_str = i[0]
                            cell_row_str = i[1:]
                            new_cell = get_column_letter(column_index_from_string(cell_colume_str) + col_idx)
                            new_cell = new_cell + str(cell_row_str)
                            if len(new_cell) != len(i):
                                new_cell = new_cell + ' '
                            dict[i] = new_cell
                        cell_key = ''.join(list(dict.keys()))
                        cell_value = ''.join(list(dict.values()))
                        new_col = value.translate(str.maketrans(cell_key, cell_value))
                        ws.cell(row=row, column=col, value=new_col)
                        dict = {}

                    pattern = r'[A-Z]+\d+'
                    value_result = re.findall(pattern, value)
                    for i in value_result:
                        cell_colume_str = i[0]
                        cell_row_str = i[1:]
                        new_cell = str(int(cell_row_str) + 1)
                        new_cell = str(cell_colume_str) + new_cell
                        dict[i] = new_cell
                    for i in dict.keys():
                        value = value.replace(i, dict[i])
                    dict = {}

            else:
                for row_idx, row in enumerate(range(start_row, end_row + 1)):
                    for col_idx, col in enumerate(range(start_column, end_column + 1)):
                        ws.cell(row=row, column=col, value=value)
    else:
        for row_idx, row in enumerate(range(start_row, end_row + 1)):
            for col_idx, col in enumerate(range(start_column, end_column + 1)):
                ws.cell(row=row, column=col, value=value)

    wb.save(file_name)


def find_first_last_data_row(file_name: str, sheet_name: [str, int], skip: int, header: bool = True,
                             first_row_offset: int = 0, last_row_offset: int = 0) -> int:
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    data = []

    for row in ws.iter_rows(values_only=True):
        if all(value is None for value in row):
            skip -= 1
            if skip == -1:
                break
        data.append(row)

    if header:
        df = pd.DataFrame(data[1:], columns=data[0])
    else:
        df = pd.DataFrame(data[1:], columns=None)

    while True:

        last_row_is_nan = df.iloc[-1].isna().all()
        if last_row_is_nan:
            df = df.drop(df.index[-1])
        else:
            break

    first_row_index = df.iloc[0].name - first_row_offset + 2
    last_row_index = df.iloc[-1].name - last_row_offset + 2

    if last_row_index < -1:
        last_row_index = -1

    return first_row_index, last_row_index


def find_replace_value(file_name: str, sheet_name: [str, int], find_values=[str, int],
                       replace_values: [str, int] = None) -> list:
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    data = []

    for row in ws.iter_rows(values_only=True):
        data.append(row)

    df = pd.DataFrame(data[1:], columns=data[0])

    indices = np.where(df.values == find_values)
    rows, cols = indices[0], indices[1]
    print(type(rows))
    result_list = []

    n = 0
    for row in rows:
        row = int(row) + 2
        col = cols[n]
        col = int(col) + 1
        result_list.append(get_column_letter(col) + str(row))
        n += 1

    if replace_values is not None:

        for i in result_list:
            ws[i].value = replace_values

        wb.save(file_name)

    return result_list


def for_each_excel_row(file_name: str) -> list:
    wb = load_workbook(file_name)
    sheet_names = wb.sheetnames

    return sheet_names

def insert_column(file_name: str, sheet_name: [str, int], start_position: [str, int], insert_column_value: [str ,list]):

    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    pattern = r'[a-z A-Z]+'
    result = re.findall(pattern, start_position)
    alphabet = result[0]

    pattern = r'\d+'
    result = re.findall(pattern, start_position)
    number = int(result[0])

    start_column = column_index_from_string(alphabet)

    if type(insert_column_value) == list:
        ws.insert_cols(start_column, amount=len(insert_column_value))
        for column_value in insert_column_value:
            ws.cell(row=number, column=start_column, value=column_value)
            start_column += 1
    else:
        ws.insert_cols(start_column)
        ws.cell(row=number, column=start_column, value=insert_column_value)
    wb.save(file_name)

def insert_row(file_name: str, sheet_name: [str, int], start_row: int, num_rows_to_insert: int):

    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    ws.insert_rows(start_row, amount=num_rows_to_insert)
    wb.save(file_name)

if __name__ == '__main__':
    insert_row('/Users/kimjunghoo/Desktop/uipath_python/연습용.xlsx', 'Sheet3', 2, 1)

